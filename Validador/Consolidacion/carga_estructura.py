import os
from openpyxl import load_workbook, Workbook

def cargar_estructura(ruta_archivo_unificado, ruta_libro_base, hoja_base, MAPEO, fila_encabezados=8):
    """
    Carga datos desde un archivo unificado a una plantilla base
    
    Args:
        ruta_archivo_unificado: Archivo origen con datos
        ruta_libro_base: Plantilla destino
        hoja_base: Nombre de la hoja en el destino
        MAPEO: Diccionario con mapeo de columnas origen -> destino
        fila_encabezados: Fila donde están los encabezados en el destino (default 8)
    """
    
    print("\n" + "="*50)
    print(" CARGANDO ESTRUCTURA")
    print("="*50)
    
    # Verificar si el archivo destino existe, si no, crearlo
    if not os.path.exists(ruta_libro_base):
        print(f"   El archivo destino no existe. Creando: {ruta_libro_base}")
        wb_new = Workbook()
        ws_new = wb_new.active
        ws_new.title = hoja_base
        wb_new.save(ruta_libro_base)
        wb_new.close()
    
    # 1. Abrir archivo origen
    wb_original = load_workbook(ruta_archivo_unificado, read_only=True, data_only=True)
    ws_original = wb_original.active
    
    # 2. Abrir archivo destino
    wb_base = load_workbook(ruta_libro_base)
    
    # Verificar que la hoja existe
    if hoja_base not in wb_base.sheetnames:
        raise Exception(f"La hoja '{hoja_base}' no existe en {ruta_libro_base}")
    
    ws_base = wb_base[hoja_base]
    
    # 3. Leer encabezados del origen (siempre fila 1)
    encabezados_origen = []
    for cell in ws_original[1]:
        encabezados_origen.append(str(cell.value or "").strip())
    
    print(f"Encabezados ORIGEN ({len(encabezados_origen)}): {encabezados_origen[:5]}...")
    
    # 4. Leer encabezados del destino (fila especificada)
    encabezados_destino = []
    # Obtener la fila específica (openpyxl usa 1-indexed)
    for cell in ws_base[fila_encabezados]:
        encabezados_destino.append(str(cell.value or "").strip())
    
    print(f"Encabezados DESTINO (fila {fila_encabezados}): {encabezados_destino[:5]}...")
    
    # 5. Verificar que hay encabezados
    if not any(encabezados_destino):
        raise Exception(f"No se encontraron encabezados en la fila {fila_encabezados} de {ruta_libro_base}")
    
    # 6. Validar mapeos
    mapeos_validos = 0
    for col_o, col_d in MAPEO.items():
        if col_o in encabezados_origen and col_d in encabezados_destino:
            mapeos_validos += 1
            print(f"   ✅ '{col_o}' → '{col_d}'")
        else:
            if col_o not in encabezados_origen:
                print(f"   Columna origen '{col_o}' no encontrada")
            if col_d not in encabezados_destino:
                print(f"   Columna destino '{col_d}' no encontrada en fila {fila_encabezados}")
    
    if mapeos_validos == 0:
        raise Exception("✖️ Ningún mapeo es válido. Verifica los nombres de las columnas.")
    
    # 7. Encontrar primera fila libre (después de la fila de encabezados)
    fila_inicio = fila_encabezados + 1
    fila_libre = fila_inicio
    
    # Buscar si ya hay datos en la primera fila de datos
    while ws_base.cell(row=fila_libre, column=1).value is not None:
        fila_libre += 1
    
    print(f"Primera fila libre: {fila_libre}")
    
    # 8. Copiar datos
    copiadas = 0
    filas_procesadas = 0
    
    for row in ws_original.iter_rows(min_row=2, values_only=True):
        filas_procesadas += 1
        datos_copiados = False
        
        for col_o, col_d in MAPEO.items():
            if col_o in encabezados_origen and col_d in encabezados_destino:
                idx_origen = encabezados_origen.index(col_o)
                idx_destino = encabezados_destino.index(col_d)
                
                valor = row[idx_origen] if idx_origen < len(row) else None
                
                if valor is not None and valor != "":
                    ws_base.cell(row=fila_libre, column=idx_destino + 1, value=valor)
                    datos_copiados = True
        
        if datos_copiados:
            copiadas += 1
            fila_libre += 1
        
        # Mostrar progreso cada 100 filas
        if filas_procesadas % 100 == 0:
            print(f"   Procesando fila {filas_procesadas}...")
    
    # 9. Guardar
    wb_base.save(ruta_libro_base)
    wb_original.close()
    wb_base.close()
    
    print(f"\nRESULTADO:")
    print(f"   Filas procesadas desde origen: {filas_procesadas}")
    print(f"   Filas copiadas al destino: {copiadas}")
    print(f"   Archivo guardado: {ruta_libro_base}")
    print("="*50)