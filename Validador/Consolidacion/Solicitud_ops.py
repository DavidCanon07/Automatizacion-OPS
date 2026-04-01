from openpyxl import load_workbook
from Configuracion_parametros import escribir
import win32com.client as win32
import pandas as pd

def obtener_solicitud_ops(origen_path, destino_path, hoja_origen, hoja_destino, mapeo_celdas, fecha_actual):
    
    # 1. Abrir y recalcular con Excel
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    excel.Visible = False
    excel.DisplayAlerts = False
    
    wb_excel = excel.Workbooks.Open(origen_path)
    excel.CalculateUntilAsyncQueriesDone()
    excel.CalculateFull()
    wb_excel.Save()  # Guarda los valores calculados
    wb_excel.Close()
    excel.Quit()
    
    # 2. Ahora abrir con openpyxl con data_only=True para obtener los valores
    wb_origen = load_workbook(origen_path, data_only=True)
    ws_origen = wb_origen[hoja_origen]
    
    # 3. Abrir destino
    wb_destino = load_workbook(destino_path)
    ws_destino = wb_destino[hoja_destino]
    
    # 4. Colocar fecha
    ws_destino["D11"].value = fecha_actual
    escribir(f"✔️ Fecha en D11: {fecha_actual}")
    
    # 5. Copiar valores calculados
    copiadas = 0
    for celda_o, celda_d in mapeo_celdas.items():
        valor = ws_origen[celda_o].value  # Ahora es el valor calculado
        
        if valor is not None and valor != "":
            ws_destino[celda_d].value = valor
            copiadas += 1
            escribir(f"✔️ {celda_o} → {celda_d}: {valor}")
        else:
            escribir(f"◬ {celda_o} está vacío")
    
    # 6. Guardar
    wb_destino.save(destino_path)
    wb_origen.close()
    wb_destino.close()
    
    escribir(f"\n✔️ Archivo guardado: {destino_path}")
    escribir(f" Celdas copiadas: {copiadas}")
    
#---------------------------------------------------------------------------------

def extraer_cuentas_y_descripciones(origen_path, destino_path, hoja_origen, hoja_destino, 
                                    celda_descripcion, celda_cuentas, celda_codigo_MIR, 
                                    columna_inicio, fila_inicio, separador, separador_codigo_MIR):
    """
    Extrae datos de columnas J, K y N, elimina duplicados y los concatena en celdas específicas.
    
    Args:
        celda_descripcion: Celda destino para las descripciones (columna J)
        celda_cuentas: Celda destino para las cuentas contables (columna K)
        celda_codigo_MIR: Celda destino para los códigos MIR (columna N)
    """
    try:
        # 1. Abrir origen con Excel
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        excel.Visible = False
        excel.DisplayAlerts = False
        wb_origen = excel.Workbooks.Open(origen_path)
        
        # Seleccionar hoja de origen
        if hoja_origen:
            ws_origen = wb_origen.Worksheets(hoja_origen)
        else:
            ws_origen = wb_origen.ActiveSheet
        
        # 2. Encontrar última fila
        ultima_fila = ws_origen.Cells(ws_origen.Rows.Count, columna_inicio).End(-4162).Row
        
        if ultima_fila < fila_inicio:
            escribir(f"◬ No hay datos en el rango especificado (fila {fila_inicio} a {ultima_fila})")
            wb_origen.Close(SaveChanges=False)
            excel.Quit()
            return []
        
        # 3. Extraer datos y eliminar duplicados
        descripcion_unicos = []
        cuentas_unicas = []
        codigos_MIR_unicos = []
        descripcion_vistos = set()
        cuentas_vistas = set()
        codigos_MIR_vistos = set()
        
        for fila in range(fila_inicio, ultima_fila + 1):
            descripcion = ws_origen.Cells(fila, columna_inicio).Value  # Columna J
            cuenta = ws_origen.Cells(fila, columna_inicio + 1).Value  # Columna K
            codigo_MIR = ws_origen.Cells(fila, columna_inicio + 4).Value  # Columna N
        
            
            # Convertir a string si es necesario
            if descripcion is not None:
                descripcion = str(descripcion).strip()
                if descripcion and descripcion not in descripcion_vistos:
                    descripcion_unicos.append(descripcion)
                    descripcion_vistos.add(descripcion)
            
            if cuenta is not None:
                cuenta = str(cuenta).strip()
                if cuenta and cuenta not in cuentas_vistas:
                    cuentas_unicas.append(cuenta)
                    cuentas_vistas.add(cuenta)
            
            if codigo_MIR is not None:
                codigo_MIR = str(codigo_MIR).strip()
                if codigo_MIR and codigo_MIR not in codigos_MIR_vistos:
                    codigos_MIR_unicos.append(codigo_MIR)
                    codigos_MIR_vistos.add(codigo_MIR)
        
        # 4. Cerrar Excel
        wb_origen.Close(SaveChanges=False)
        excel.Quit()
        
        # 5. Concatenar valores (puedes cambiar el separador según necesites)
        separador = separador
        separador_codigo_MIR = separador_codigo_MIR
        texto_descripciones = separador.join(descripcion_unicos)
        texto_cuentas = separador.join(cuentas_unicas)
        texto_codigos_MIR = separador_codigo_MIR.join(codigos_MIR_unicos)
        
        # 6. Abrir destino y asignar
        wb_destino = load_workbook(destino_path)
        ws_destino = wb_destino[hoja_destino]
        
        # Asignar valores a las celdas destino
        ws_destino[celda_descripcion].value = texto_descripciones
        ws_destino[celda_cuentas].value = texto_cuentas
        ws_destino[celda_codigo_MIR].value = texto_codigos_MIR
        
        # Guardar y cerrar
        wb_destino.save(destino_path)
        wb_destino.close()
        
        escribir(f"\n✔️ Datos asignados:")
        escribir(f"   {celda_descripcion} (Descripciones): {len(descripcion_unicos)} únicos → \n{texto_descripciones}")
        escribir(f"   {celda_cuentas} (Cuentas Contables): {len(cuentas_unicas)} únicas → \n {texto_cuentas}")
        escribir(f"   {celda_codigo_MIR} (Códigos MIR): {len(codigos_MIR_unicos)} únicos → \n {texto_codigos_MIR}")
        
        return {"descripciones": descripcion_unicos, "cuentas": cuentas_unicas, "codigos_MIR": codigos_MIR_unicos}
        
    except Exception as e:
        escribir(f"✖️ Error: {e}")
        return {}
    

# generar archivo plano de OPS DDMMYYYY
def exportar_txt_limpio(ruta_formato_ops, hoja, rango, archivo_txt):
    """
    Exporta directamente a txt SIN líneas vacías al final
    """
    import pandas as pd
    
    # Leer datos
    df = pd.read_excel(
        ruta_formato_ops, 
        skiprows=8,
        sheet_name=hoja,
        header=None,
        usecols=rango
    )
    # 1. Eliminar filas completamente vacías (NaN)
    df = df.dropna(how='all')
    
    # 2. Eliminar filas donde todos los valores son strings vacíos
    df = df[df.apply(lambda row: row.astype(str).str.strip().ne('').any(), axis=1)]
    
    # 3. Eliminar filas donde el primer valor está vacío o es '0'
    if len(df.columns) > 0:
        df = df[df.iloc[:, 0].astype(str).str.strip().notna()]
        df = df[df.iloc[:, 0].astype(str).str.strip() != '']
        df = df[~df.iloc[:, 0].astype(str).str.strip().isin(['0', '0.0', 'nan', 'NaN'])]
    
    print(f" Filas después de limpiar: {len(df)}")
    
    # 4. Exportar línea por línea sin salto final
    with open(archivo_txt, 'w', encoding='utf-8') as f:
        for i, row in df.iterrows():
            # Unir todos los valores de la fila
            valores = [str(val).strip() for val in row.values if pd.notna(val)]
            linea = ''.join(valores)  # Sin separador
            
            # Escribir línea con salto excepto la última
            if i < len(df) - 1:
                f.write(linea + '\n')
            else:
                f.write(linea)
    
    # Verificar el archivo generado
    with open(archivo_txt, 'r', encoding='utf-8') as f:
        lineas_finales = f.readlines()
    
    print(f" Líneas finales en archivo: {len(lineas_finales)}")
    
    escribir(f"✔️ Archivo plano generado: {archivo_txt}")
    escribir(f"   Total líneas: {len(lineas_finales)}")
    
    return len(lineas_finales)

