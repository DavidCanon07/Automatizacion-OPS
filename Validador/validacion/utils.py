import re, os
import time as t
from openpyxl.utils import get_column_letter
import shutil
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

#-------------------------------------------------------------------------------------------------------------
#Función para convertir nombres de columnas extra a un formato más legible
def formatea_nombre_columna(nombre: str) -> str:
    
    #Si el nombre es 'Unnamed: N', devuelve 'Columna {LETRA}' (A=1).
    #Si no, devuelve el nombre original.
    #Ejemplo: 'Unnamed: 0' -> 'Columna A', 'Unnamed: 1' -> 'Columna B', etc.
    m = re.match(r"^Unnamed:\s*(\d+)$", str(nombre), flags=re.IGNORECASE)
    if m:
        idx0 = int(m.group(1))           # índice 0-based
        letra = get_column_letter(idx0 + 1)  # Excel es 1-based
        return f"Columna {letra}"
    return str(nombre)

#-------------------------------------------------------------------------------------------------------------


#Función para imprimir texto con efecto de máquina de escribir
def escribir(texto, velocidad=0.01):
    for char in texto:
        print(char, end="", flush=True)
        t.sleep(velocidad)
    print()  # Salto de línea solo al final
    


def input_con_efecto(texto, velocidad=0.01):
    """Muestra texto con efecto máquina de escribir y retorna input del usuario"""
    for char in texto:
        print(char, end="", flush=True)
        t.sleep(velocidad)
    print()  # Salto de línea
    return input()  # input sin prompt


#Función para crear carpetas si no existen
def crear_carpeta_si_no_existe(base_path, carpeta_archivos, ejecucion, formatos, estructura, historico):
    try:
        for path in [base_path, carpeta_archivos, ejecucion, formatos, estructura, historico]:
            if not path.exists():
                escribir(f"Creando carpeta: {path}")
                path.mkdir(parents=True, exist_ok=True)
            else:
                continue
    except Exception as e:
        escribir(f"Error al crear carpetas: {e}")
        
#Función para validar que las estructuras necesarias para la consolidación existan en la carpeta de estructuras
def validar_estructuras(ruta_libro_base, ruta_libro_base_debitos, ruta_libro_base_solicitud, ruta_consolidado_historico,
                        formato_ops, formato_debitos, formato_solicitud, archivo_historico):
    try:
        #convertir a Path si no lo son
        ruta_libro_base = Path(ruta_libro_base)
        ruta_libro_base_debitos = Path(ruta_libro_base_debitos)
        ruta_libro_base_solicitud = Path(ruta_libro_base_solicitud)
        ruta_consolidado_historico = Path(ruta_consolidado_historico)
        
        # Verificar cada estructura
        if not ruta_libro_base.is_file():
            escribir(f"Copiando estructura base: {ruta_libro_base.name}")
            shutil.copy2(formato_ops, ruta_libro_base)
        
        if not ruta_libro_base_debitos.is_file():
            escribir(f"Copiando estructura débitos: {ruta_libro_base_debitos.name}")
            shutil.copy2(formato_debitos, ruta_libro_base_debitos)
        
        if not ruta_libro_base_solicitud.is_file():
            escribir(f"Copiando estructura solicitud: {ruta_libro_base_solicitud.name}")
            shutil.copy2(formato_solicitud, ruta_libro_base_solicitud)
            
        if not ruta_consolidado_historico.is_file():
            escribir(f"Copiando archivo consolidado historico: {ruta_consolidado_historico.name}")
            shutil.copy2(archivo_historico, ruta_consolidado_historico)
            
            
    except Exception as e:
        escribir(f"Error al validar estructuras: {e}")
        

def formato_para_columnas(archivo, hoja, columnas_formato):
    
    if not os.path.exists(archivo):
        escribir(f"✖️ El archivo duplicados no se encuentra: {archivo}")
        return None
    
    try:
        wb = load_workbook(archivo)
        ws = wb[hoja]
        
        # Definir estilos
        estilo_encabezado = Font(bold=True, color='FFFFFF')
        fondo_encabezado = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        
        estilo_datos = Font(bold=False)
        fondo_azul_claro = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
        fondo_normal = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        
        centrado = Alignment(horizontal='center', vertical='center')
        
        # Si columnas_formato es None, usar todas las columnas del archivo
        if columnas_formato is None:
            # Obtener nombres de columnas desde el encabezado
            columnas = []
            for col in range(1, ws.max_column + 1):
                nombre_col = ws.cell(row=1, column=col).value
                if nombre_col:
                    columnas.append(nombre_col)
        else:
            columnas = columnas_formato
        
        # Formatear SOLO los encabezados de interes el indice
        indices_columnas = []
        for col in range(1, ws.max_column + 1):
            nombre_col = ws.cell(row=1, column=col).value
            if nombre_col in columnas:
                indices_columnas.append(col)
        
        # Formatear SOLO esos encabezados
        for col_idx in indices_columnas:
            celda = ws.cell(row=1, column=col_idx)
            celda.font = estilo_encabezado
            celda.fill = fondo_encabezado
            celda.alignment = centrado
        
        # Crear diccionario para saber qué columna es cada nombre
        columna_indices = {}
        for col in range(1, ws.max_column + 1):
            nombre = ws.cell(row=1, column=col).value
            if nombre:
                columna_indices[nombre] = col
        
        # Formatear usando los índices correctos
        for col_name in columnas:
            if col_name in columna_indices:
                col_idx = columna_indices[col_name]
                letra_col = chr(64 + col_idx)
                
                # Calcular ancho máximo
                max_len = len(str(col_name))
                for fila in range(2, ws.max_row + 1):
                    valor = ws[f"{letra_col}{fila}"].value
                    if valor:
                        max_len = max(max_len, len(str(valor)))
                
                ancho = min(max_len + 2, 50)
                ws.column_dimensions[letra_col].width = ancho
                
                # Aplicar formato a la columna
                for fila in range(2, ws.max_row + 1):
                    celda = ws[f"{letra_col}{fila}"]
                    celda.font = Font(bold=True)
                    celda.fill = fondo_azul_claro
                    celda.alignment = centrado
            else:
                escribir(f"⚠️ Columna '{col_name}' no encontrada en el archivo")
        
        wb.save(archivo)
        escribir(f"✔️ Archivo exportado y formateado: {archivo}")
        escribir(f"   Columnas formateadas: {columnas}")
        
        return True
    
    except Exception as e:
        escribir(f"✖️ Error al abrir el archivo duplicados: {e}")
        return None