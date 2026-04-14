import pandas as pd
from validacion.utils import escribir
from validacion.reportes import exportar_errores
import os
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime, timedelta


#funcion para comparar dataframes
def comparar_consolidados(ruta_historico, ruta_archivo_unificado, ruta_retorno_duplicados, sheet):
    """
    Compara el archivo unificado contra el histórico para encontrar duplicados
    
    Args:
        ruta_historico: Ruta del archivo histórico (base)
        ruta_archivo_unificado: Ruta del archivo nuevo a validar
        ruta_retorno_duplicados: Ruta donde guardar el reporte de duplicados
    """
    ruta_retorno_duplicados = Path(ruta_retorno_duplicados)
    
    # Verificar que los archivos existen
    if not os.path.exists(ruta_historico):
        escribir(f"✖️ El archivo histórico no se encuentra: {ruta_historico}")
        return None
    
    if not os.path.exists(ruta_archivo_unificado):
        escribir(f"✖️ El archivo unificado no se encuentra: {ruta_archivo_unificado}")
        return None
    
    try:
        # Leer los DataFrames
        df_historico = pd.read_excel(ruta_historico, sheet_name=sheet)
        df_nuevo = pd.read_excel(ruta_archivo_unificado)
        
        # Verificar que la columna 'Concatenado' existe
        if 'Concatenado' not in df_historico.columns:
            escribir(f" La columna 'Concatenado' no existe en el archivo histórico")
            # Crear columna concatenado si no existe
            df_historico['Concatenado'] = df_historico.astype(str).agg(''.join, axis=1)
        
        if 'Concatenado' not in df_nuevo.columns:
            escribir(f" La columna 'Concatenado' no existe en el archivo nuevo")
            df_nuevo['Concatenado'] = df_nuevo.astype(str).agg(''.join, axis=1)
        
        # Encontrar registros del NUEVO archivo que ya existen en el HISTÓRICO
        duplicados = df_nuevo[df_nuevo['Concatenado'].isin(df_historico['Concatenado'])]
        
        # Resultados
        if not duplicados.empty:
            escribir(f" Se encontraron {len(duplicados)} registros duplicados")
            
            # Agregar columnas informativas
            duplicados = duplicados.copy()
            duplicados['Número de fila registrada en historico'] = df_historico['N°']
            duplicados['Fecha_historico'] = df_historico['FECHA'].astype(str)
            duplicados['descripcion'] = "REGISTRO YA EXISTE EN EL ARCHIVO HISTÓRICO"
            
            # Exportar a Excel
            exportar_errores(
                duplicados,
                ruta_retorno_duplicados,
                "⚠ Se encontraron duplicados en el archivo comparado con el histórico. "
                "Revisar archivo: ALERTA_DUPLICADOS_HISTÓRICO.xlsx",
                sheet_name="Alerta Duplicados",
                index=False
            )
            
            return duplicados
        else:
            escribir("✔️ Validación de duplicados completada sin alertas.")
            return None
            
    except Exception as e:
        escribir(f"✖️ Error al comparar DataFrames: {e}")
        return None
    


def depurar_historico(ruta_archivo, columna_fecha, hoja, dias_ventana):
    """
    Mantiene solo los registros de los últimos N días usando openpyxl
    Elimina filas directamente sin reescribir todo el archivo
    """
    try:
        # Cargar el libro
        wb = load_workbook(ruta_archivo)
        ws = wb[hoja]
        
        # Encontrar la columna de fecha y la última fila con datos
        fecha_col_idx = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == columna_fecha:
                fecha_col_idx = col
                break
        
        if fecha_col_idx is None:
            escribir(f"✖️ Columna '{columna_fecha}' no encontrada")
            return None
        
        # Fecha límite
        fecha_limite = datetime.now() - timedelta(days=dias_ventana)
        fecha_limite_str = fecha_limite.strftime('%Y-%m-%d')
        
        # Recopilar filas a eliminar (desde la última fila hacia arriba)
        filas_a_eliminar = []
        registros_totales = 0
        registros_mantenidos = 0
        
        for row in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=fecha_col_idx).value
            registros_totales += 1
            
            # Convertir a fecha si es necesario
            if cell_value:
                if isinstance(cell_value, datetime):
                    fecha_fila = cell_value.date()
                elif isinstance(cell_value, str):
                    try:
                        fecha_fila = datetime.strptime(cell_value, '%d/%m/%Y').date()
                    except:
                        fecha_fila = None
                else:
                    fecha_fila = None
                
                if fecha_fila and fecha_fila >= fecha_limite.date():
                    registros_mantenidos += 1
                else:
                    filas_a_eliminar.append(row)
        
        # Eliminar filas (de abajo hacia arriba para no afectar índices)
        for row in reversed(filas_a_eliminar):
            ws.delete_rows(row)
        
        # Guardar
        wb.save(ruta_archivo)
        wb.close()
        
        registros_eliminados = registros_totales - registros_mantenidos
        
        escribir(f"\n Depuración de histórico completada:")
        escribir(f"   Ventana: {dias_ventana} días")
        escribir(f"   Fecha límite: {fecha_limite.strftime('%d/%m/%Y')}")
        escribir(f"   Registros originales: {registros_totales}")
        escribir(f"   Registros mantenidos: {registros_mantenidos}")
        escribir(f"   Registros eliminados: {registros_eliminados}")
        escribir(f"✔️ Histórico depurado y guardado: {ruta_archivo}")
        
        return True
        
    except Exception as e:
        escribir(f"✖️ Error al depurar histórico: {e}")
        return None

